# https://www.scottbrady91.com/Email-Verification/Python-Email-Verification-Script
# https://pypi.python.org/pypi/validate_email
import pandas as pd
import datetime
import openpyxl
from openpyxl.styles import PatternFill
from Shared.file_service import FileService
from Shared.common import Common
from Shared.enums import FileType, WorkSheet as WS, PATH as p, FilePath as pth
import warnings
import os
from dateutil import parser


class BapQA:
	def __init__(self):
		box_path = Common.change_location(p.DATA)

		fl = FileService(box_path)
		self.ric_files = fl.get_source_file()

		self.okay = PatternFill(fgColor='E1F7DC', bgColor='C00000', fill_type='solid')
		self.amber = PatternFill(fgColor='F4B042', bgColor='C00000', fill_type='solid')
		self.header = PatternFill(fgColor='218c04', bgColor='C00000', fill_type='solid')
		self.empty = PatternFill(fgColor='f9462a', bgColor='C00000', fill_type='solid')
		self.red = PatternFill(fgColor='f72f11', bgColor='C00000', fill_type='solid')

		warnings.filterwarnings("ignore")

		self.quarter = 'Q1'
		self.year = '2019'
		self.youth = 'Youth'
		self.all_youth = 'ALL incl. youth'

		self.month_names = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']
		self.no_value = ['na', 'n/a', '','0000-00-00', '*****']

	def proper_stage(self, stg):
		stage = ['discovery', 'efficiency', 'idea', 'scale', 'validation']
		val = [s for s in stage if s in stg.lower()]
		return len(val) > 0

	def yes_no(self, resp):
		response = ['y', 'yes', 'n', 'no']
		val = [r for r in response if r == resp.lower()]
		return len(val) > 0

	def check_columns_completeness(self):
		dfps = pd.DataFrame()
		dfpys = pd.DataFrame()
		dfqc = pd.DataFrame()
		dfac = pd.DataFrame()


		clm_lst = []
		for fl in self.ric_files:
			Common.change_location(p.DATA)
			wb = openpyxl.load_workbook(fl, data_only=True)
			ric_file_name = fl[:-5]
			print('-' * 250)

			program_sheet = wb.get_sheet_by_name(WS.bap_program.value)
			df_ps = self.sheet_columns(program_sheet, ric_file_name, WS.bap_program.value)
			program_youth_sheet = wb.get_sheet_by_name(WS.bap_program_youth.value)
			df_pys = self.sheet_columns(program_youth_sheet, ric_file_name, WS.bap_program_youth.value)
			quarterly_company_sheet = wb.get_sheet_by_name(WS.bap_company.value)
			df_qc = self.sheet_columns(quarterly_company_sheet, ric_file_name, WS.bap_company.value)
			annual_company_sheet = wb.get_sheet_by_name(WS.bap_company_annual.value)
			df_ac = self.sheet_columns(annual_company_sheet, ric_file_name, WS.bap_company_annual.value)

			dfps = pd.concat([dfps, df_ps])
			dfpys = pd.concat([dfpys, df_pys])
			dfqc = pd.concat([dfqc, df_qc])
			dfac = pd.concat([dfac, df_ac])


		writer = pd.ExcelWriter('00 ALL_RIC_BAP_COLUMNS_FY19_Q1.xlsx')
		dfps.to_excel(writer, 'Program', index=False)
		dfpys.to_excel(writer, 'Program Youth', index=False)
		dfqc.to_excel(writer, 'Quarterly Company', index=False)
		dfac.to_excel(writer, 'Annual Company', index=False)

		Common.change_location(p.QA)
		print(os.getcwd())
		writer.save()

	def check_rics_file(self, loc, dest, combined=False):
		path = Common.change_working_directory(loc.value)
		fl = FileService(path)
		self.ric_files = fl.get_source_file()
		if combined:
			self.ric_files = [f for f in self.ric_files if 'ETL_RICS_BAP_COMBINED' in f]
		print('-' * 100, '\nPROCESSING RIC FILES\n')
		for fl in self.ric_files:
			print(fl)
			try:
				Common.change_working_directory(loc.value)
				wb = openpyxl.load_workbook(fl, data_only=True)

				print('\tProgram data')
				program_sheet = wb.get_sheet_by_name(WS.bap_program.value)
				self.qa_program_sheet(program_sheet)
				print('\tProgram Youth data')
				program_youth_sheet = wb.get_sheet_by_name(WS.bap_program_youth.value)
				self.qa_program_youth_sheet(program_youth_sheet)
				print('\tQuarterly Company data')
				quarterly_company_sheet = wb.get_sheet_by_name(WS.bap_company.value)
				self.rics_sheet_header(quarterly_company_sheet)
				self.qa_quarterly_company_data_sheet(quarterly_company_sheet)
				print('\tAnnual Company data')
				if self.quarter == 'Q3' and ('haltec' in fl.lower() or 'communitech' in fl.lower()):
					annual_company_sheet = wb.get_sheet_by_name(WS.bap_company_annual.value)
					self.qa_annual_company_data_sheet(annual_company_sheet)

				Common.change_working_directory(dest.value)
				wb.save('{}_Q1_QA.xlsx'.format(fl[:-5]))
				print('\t\t[{}_Q1_QA] IS SAVED.'.format(fl[:-5]))
			except BaseException as ex:
				print(ex)

	def rics_sheet_header(self, sheet):
		for cl in sheet.columns:
			for c in cl:
				if c.row > 1:
					if c.value is None:
						c.fill = self.amber

	def qa_program_youth_sheet(self, sheet):
		for cl in sheet.columns:
			for c in cl:
				if c.row == 1:
					c.fill = self.header
				elif c.row > 1:
					if c.column in ['A', 'B', 'C', 'H']:
						if isinstance(int(c.value), int):
							c.fill = self.okay
						else:
							c.fill = self.amber
					if c.column == 'D':
						c.value = self.quarter
						c.fill = self.okay
					else:
						c.fill = self.amber
					if c.column == 'E':
						c.value = self.year
						c.fill = self.okay
					else:
						c.fill = self.amber
					if c.column == 'F':
						if str(c.value).lower() == self.youth.lower():
							c.fill = self.okay
						else:
							c.fill = self.amber

	def qa_program_sheet(self, sheet):
		for cl in sheet.columns:
			for c in cl:
				try:
					if c.row == 1:
						c.fill = self.header
					elif c.row > 1:
						if c.column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'I', 'J', 'K', 'L', 'M', 'N']:
							if isinstance(int(c.value), int):
								c.fill = self.okay
							else:
								c.fill = self.amber
						if c.column == 'O':
							c.value = self.quarter
							c.fill = self.okay
						else:
							c.fill = self.amber
						if c.column == 'P':
							c.value = self.year
							c.fill = self.okay
						else:
							c.fill = self.amber
						if c.column == 'Q':
							if c.value.lower() == self.all_youth.lower():
								c.fill = self.okay
							else:
								c.fill = self.amber
				except Exception as ex:
					c.fill = self.red
					print('{} | {} | {} | {}'.format(c.column, c.row, c.value, ex))

	def qa_quarterly_company_data_sheet(self, sheet):
		for cl in sheet.columns:
			for c in cl:
				try:
					if c.value is not None:
						if c.row == 1:
							c.fill = self.header
						elif c.row > 1:
							if c.column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'V', 'H', 'I', 'K', 'L', 'V', 'X', 'AA']:
								if len(str(c.value)) > 0:
									c.fill = self.okay
								else:
									c.fill = self.amber
							if c.column == 'J':
								if self.proper_stage(c.value):
									c.fill = self.okay
								else:
									c.fill = self.amber
							if c.column in ['AD', 'AF', 'AH','M', 'Q', 'R', 'S','T', 'U', 'Y', 'Z', 'AI', 'AJ']:
								if isinstance(c.value, float) or isinstance(c.value, int):
									c.fill = self.okay
								else:
									c.fill = self.amber
							if c.column == 'O':
								if isinstance(c.value, str):
									if str(c.value).lower() in self.no_value:
										c.fill = self.amber
								elif isinstance(parser.parse(str(c.value)), datetime.date) or c.value.lower() in self.month_names:
									c.fill = self.okay
								elif isinstance(int(c.value), int) and (int(c.value) > 0 and int(c.value) < 13):
									c.fill = self.okay
								elif isinstance(int(c.value), int) and int(c.value) == 0:
									c.fill = self.amber
								else:
									c.fill = self.amber
							if c.column == 'P':
								if str(c.value).lower() in self.no_value:
									c.fill = self.amber
								elif isinstance(parser.parse(str(c.value)), datetime.date) or (isinstance(int(c.value), int) and int(c.value) > 1000):
									c.fill = self.okay
								elif isinstance(int(c.value), int) and int(c.value) == 0:
									c.fill = self.amber
								else:
									c.fill = self.amber
							if c.column in ['N', 'V', 'AB', 'AA', 'AC']:
								if str(c.value) == '*****':
									c.fill = self.amber
								elif self.yes_no(c.value):
									c.fill = self.okay
								else:
									c.fill = self.amber
							if c.column == 'W':
								if str(c.value).lower() in self.no_value:
									c.fill = self.amber
								elif isinstance(parser.parse(str(c.value)), datetime.datetime):
									c.fill = self.okay
								else:
									c.fill = self.amber
							if c.column == 'AK':
								if str(c.value) == self.quarter:
									c.fill = self.okay
								else:
									c.fill = self.amber
							if c.column == 'AL':
								if str(c.value) == self.year:
									c.fill = self.okay
								else:
									c.fill = self.amber
				except Exception as ex:
					c.fill = self.red
					print('{} | {} | {} | {}'.format(c.column, c.row, c.value, ex))

	def qa_annual_company_data_sheet(self, sheet):
		for cl in sheet.columns:
			for c in cl:
				# if c.value is not None:
				try:
					if c.row == 1:
						c.fill = self.header
					elif c.row > 1:
						if c.column in ['A', 'B', 'C', 'D']:
							if len(c.value) > 0:
								c.fill = self.okay
							else:
								c.fill = self.amber
						if c.column in ['E', 'F', 'K']:
							if isinstance(c.value, float):
								c.fill = self.okay
							else:
								c.fill = self.amber
						if c.column in ['G', 'H', 'I', 'J']:
							if isinstance(c.value, int):
								c.fill = self.okay
							else:
								c.fill = self.amber
						if c.column == 'L':
							c.value = '--'
						if c.column == 'M':
							c.value = '2019'
				except Exception as ex:
					c.fill = self.red
					print('{} | {} | {} | {}'.format(c.column, c.row, c.value, ex))
				# else:
				# 	c.fill = self.amber

	def sheet_columns(self, sheet, ric, sheet_name):
		lst = []
		print('{} - {}'.format(ric.upper(), sheet_name))
		for rw in sheet.rows:
			for i in range(len(rw)):
				if rw[i].row == 1:
					d = dict()
					d['RIC'] = ric
					d['Sheet'] = sheet_name
					d['Letter'] = rw[i].column
					d['Header'] = rw[i].value
					lst.append(d)
		return pd.DataFrame(lst, columns=lst[0].keys())


if __name__ == '__main__':
	bapqa = BapQA()
	bapqa.proper_stage('Stage 0 - Discovery')